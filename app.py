"""
Sistema de Gestión Escolar (SGE) - Universidad de Enfermería
--------------------------------------------------------------
Paso 1: Configuración inicial de Flask + Modelos de Base de Datos (SQLAlchemy)

Reglas de negocio implementadas a nivel de modelo:
  1. Los docentes NO tienen tabla/rol de acceso: no existe modelo "Maestro".
  2. "Escudo" del Plan de Estudios: Materia siempre pertenece a un PlanEstudio,
     y Alumno siempre pertenece a un PlanEstudio. La validación de que una
     calificación solo pueda capturarse si la Materia pertenece al Plan del
     Alumno se hace a nivel de lógica de aplicación (Paso 4 - Módulo de
     Captura de Calificaciones), pero el modelo ya deja la estructura lista
     mediante las relaciones y un método de validación auxiliar.
  3. Centrado en el Alumno: Calificacion se vincula a matricula (Alumno),
     nunca a un "Grupo". El alumno puede cambiar de grupo/generación sin
     perder su historial porque el historial cuelga de su matrícula.
"""

import enum
import io
import os
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from functools import wraps

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from urllib.parse import urlparse

from flask import (
    Flask, render_template, request, flash, redirect, url_for, abort,
    session, send_file, send_from_directory
)
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user
)
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_mail import Mail, Message
from sqlalchemy import UniqueConstraint, CheckConstraint, or_, func
from sqlalchemy.exc import IntegrityError
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from markupsafe import Markup, escape

from config import config_by_name

db = SQLAlchemy()
migrate = Migrate()
login_manager = LoginManager()
csrf = CSRFProtect()
limiter = Limiter(key_func=get_remote_address)
mail = Mail()


# ---------------------------------------------------------------------------
# ENUMS
# ---------------------------------------------------------------------------

class EstatusAlumno(enum.Enum):
    """
    Estatus del ciclo de vida del alumno dentro del sistema.
    'PENDIENTE' es el estatus inicial obligatorio al llegar desde el
    Módulo de Auto-registro Público (Módulo 1).
    """
    PENDIENTE = 'Pendiente de Validación'
    ACTIVO = 'Activo'
    BAJA_TEMPORAL = 'Baja Temporal'
    BAJA_DEFINITIVA = 'Baja Definitiva'
    EGRESADO = 'Egresado'


class TurnoAlumno(enum.Enum):
    MATUTINO = 'Matutino'
    VESPERTINO = 'Vespertino'
    MIXTO = 'Mixto'


class ModalidadEstudio(enum.Enum):
    ESCOLARIZADO = 'Escolarizado'
    SEMIESCOLARIZADO = 'Semiescolarizado'
    DISTANCIA = 'A Distancia'


class RolUsuario(enum.Enum):
    """
    Roles humanos del sistema; los docentes NO tienen rol (regla de
    negocio: cero acceso a maestros).

      - DIRECTIVO: control total. Único que crea/gestiona cuentas y borra
        documentos.
      - ADMINISTRATIVO: acceso a TODO (alumnos + cobros) salvo lo
        exclusivo de Dirección (borrar documentos, usuarios).
      - CONTADOR: acceso COMPLETO al área de Cobros (crear/cancelar
        cargos, catálogo de conceptos, configurar recargos). Sin acceso
        al área de administración de alumnos.
      - CAPTURADOR: acceso a la administración académica del alumno
        (documentos, estatus, boletas, historial). Sin acceso a Cobros.
    """
    DIRECTIVO = 'Dirección / Administrador General'
    ADMINISTRATIVO = 'Administrativo (Alumnos + Cobros)'
    CONTADOR = 'Contador (Área de Cobros)'
    CAPTURADOR = 'Capturador (Administración del Alumno)'


# ---------------------------------------------------------------------------
# USUARIOS DEL SISTEMA (login)
# ---------------------------------------------------------------------------

class Usuario(UserMixin, db.Model):
    """
    Cuenta de acceso al sistema. UserMixin le da a Flask-Login las
    propiedades que necesita (is_authenticated, is_active, get_id, etc.).
    La contraseña NUNCA se guarda en texto plano: solo su hash.
    """
    __tablename__ = 'usuarios'

    id = db.Column(db.Integer, primary_key=True)
    nombre_completo = db.Column(db.String(150), nullable=False)
    username = db.Column(db.String(50), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    rol = db.Column(db.Enum(RolUsuario), nullable=False, default=RolUsuario.ADMINISTRATIVO)
    activo = db.Column(db.Boolean, default=True, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    ultimo_acceso = db.Column(db.DateTime, nullable=True)

    def set_password(self, password_plano: str) -> None:
        # pbkdf2:sha256 (default de Werkzeug): estándar robusto y ampliamente auditado.
        self.password_hash = generate_password_hash(password_plano)

    def check_password(self, password_plano: str) -> bool:
        return check_password_hash(self.password_hash, password_plano)

    def es_directivo(self) -> bool:
        return self.rol == RolUsuario.DIRECTIVO

    def puede_cobros(self) -> bool:
        return self.rol in (RolUsuario.DIRECTIVO, RolUsuario.ADMINISTRATIVO, RolUsuario.CONTADOR)

    def puede_alumnos(self) -> bool:
        return self.rol in (RolUsuario.DIRECTIVO, RolUsuario.ADMINISTRATIVO, RolUsuario.CAPTURADOR)

    @property
    def is_active(self):
        # Sobreescribe el default de UserMixin (que siempre es True):
        # una cuenta desactivada por Dirección no puede iniciar sesión.
        return self.activo

    def __repr__(self):
        return f'<Usuario {self.username} ({self.rol.name})>'


# ---------------------------------------------------------------------------
# MODELOS
# ---------------------------------------------------------------------------

class PlanEstudio(db.Model):
    """
    Representa un Plan de Estudios oficial (ej. "Licenciatura en Enfermería
    Generación 2024"). Es el contenedor "blindado" de materias: nada se
    agrega aquí desde la captura de calificaciones, solo desde administración
    del plan.
    """
    __tablename__ = 'planes_estudio'

    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(150), nullable=False)
    clave_carrera = db.Column(db.String(10), nullable=False)  # Ej: "LEN" -> usado para folio de matrícula
    anio_generacion = db.Column(db.Integer, nullable=False)
    duracion_anios = db.Column(db.Integer, nullable=True)  # Ej. 3 (para mostrar en la Ficha de Inscripción)
    monto_mensualidad = db.Column(db.Numeric(10, 2), nullable=True)  # Precio de mensualidad de esta carrera (solo Directivo lo edita)
    activo = db.Column(db.Boolean, default=True, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

    # Relaciones
    materias = db.relationship(
        'Materia',
        backref='plan',
        lazy=True,
        cascade='all, delete-orphan'
    )
    alumnos = db.relationship(
        'Alumno',
        backref='plan',
        lazy=True
    )

    __table_args__ = (
        UniqueConstraint('clave_carrera', 'anio_generacion', name='uq_plan_clave_anio'),
    )

    def __repr__(self):
        return f'<PlanEstudio {self.clave_carrera}-{self.anio_generacion}: {self.nombre}>'


class Materia(db.Model):
    """
    Materia perteneciente OBLIGATORIAMENTE a un PlanEstudio.
    No puede existir una Materia "suelta": id_plan_fk es NOT NULL.
    Esto es lo que garantiza el "Escudo" del Plan de Estudios: la
    aplicación solo debe permitir seleccionar materias filtradas por
    Materia.query.filter_by(id_plan_fk=alumno.id_plan_fk).
    """
    __tablename__ = 'materias'

    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(150), nullable=False)
    clave = db.Column(db.String(20), nullable=True)  # Clave oficial de la materia, opcional
    cuatrimestre = db.Column(db.Integer, nullable=False)
    creditos = db.Column(db.Float, nullable=True)

    id_plan_fk = db.Column(
        db.Integer,
        db.ForeignKey('planes_estudio.id'),
        nullable=False
    )

    # Relaciones
    calificaciones = db.relationship(
        'Calificacion',
        backref='materia',
        lazy=True
    )

    __table_args__ = (
        CheckConstraint('cuatrimestre > 0', name='ck_materia_cuatrimestre_positivo'),
        UniqueConstraint('id_plan_fk', 'nombre', 'cuatrimestre', name='uq_materia_por_plan'),
    )

    def __repr__(self):
        return f'<Materia {self.nombre} (Cuatri {self.cuatrimestre}) - Plan {self.id_plan_fk}>'


class Alumno(db.Model):
    """
    Entidad central del sistema. La matrícula (matricula_id) es la Primary
    Key y funciona como identificador único de negocio (no un id numérico
    autoincremental interno), ya que es el dato con el que Control Escolar
    y el propio alumno identifican su expediente.
    """
    __tablename__ = 'alumnos'

    matricula_id = db.Column(db.String(20), primary_key=True)  # Ej: LEN2024-00015
    nombre_completo = db.Column(db.String(200), nullable=False)
    curp = db.Column(db.String(18), unique=True, nullable=False, index=True)
    fecha_nacimiento = db.Column(db.Date, nullable=False)
    fecha_certificado_prepa = db.Column(db.Date, nullable=False)

    id_plan_fk = db.Column(
        db.Integer,
        db.ForeignKey('planes_estudio.id'),
        nullable=False
    )

    estatus = db.Column(
        db.Enum(EstatusAlumno),
        default=EstatusAlumno.PENDIENTE,
        nullable=False
    )

    grupo_actual = db.Column(db.String(20), nullable=True)  # Informativo; NUNCA usado como FK de calificaciones
    correo = db.Column(db.String(150), nullable=True)
    telefono = db.Column(db.String(20), nullable=True)  # Teléfono fijo
    telefono_movil = db.Column(db.String(20), nullable=True)

    # --- Datos personales adicionales (para la Ficha de Inscripción impresa) ---
    sexo = db.Column(db.String(20), nullable=True)  # Femenino / Masculino / Otro
    numero_identificacion = db.Column(db.String(30), nullable=True)  # Nº de INE u otra identificación oficial
    estado_civil = db.Column(db.String(30), nullable=True)
    nacionalidad = db.Column(db.String(50), nullable=True, default='Mexicana')
    tipo_sangre = db.Column(db.String(5), nullable=True)  # Ej. "O+", "A-"

    # --- Domicilio ---
    domicilio_calle_numero = db.Column(db.String(200), nullable=True)
    domicilio_ciudad = db.Column(db.String(100), nullable=True)
    domicilio_cp = db.Column(db.String(10), nullable=True)
    domicilio_estado = db.Column(db.String(100), nullable=True)

    # --- Contacto de emergencia (persona distinta al tutor legal) ---
    contacto_emergencia_nombre = db.Column(db.String(150), nullable=True)
    contacto_emergencia_telefono = db.Column(db.String(20), nullable=True)
    contacto_emergencia_parentesco = db.Column(db.String(50), nullable=True)  # Ej. "Hermana", "Madre"

    como_se_entero = db.Column(db.String(200), nullable=True)  # Ej. "Por una amiga", "Redes sociales"

    turno = db.Column(db.Enum(TurnoAlumno), nullable=True)
    modalidad = db.Column(db.Enum(ModalidadEstudio), nullable=True)

    # --- Datos ampliados del expediente (Módulo de Documentos) ---
    escuela_prepa = db.Column(db.String(200), nullable=True)
    alergias_condiciones = db.Column(db.Text, nullable=True)
    telefono_tutor = db.Column(db.String(20), nullable=True)

    # --- Seguimiento administrativo y académico (Control Escolar) ---
    # cuatrimestre_actual: en qué cuatrimestre inició/va el alumno. Por defecto 1
    # (nuevo ingreso), pero se puede ajustar en casos de revalidación/equivalencia
    # donde el alumno entra directo a un cuatrimestre avanzado.
    cuatrimestre_actual = db.Column(db.Integer, nullable=False, default=1)
    documentacion_pendiente = db.Column(db.Text, nullable=True)  # Ej. "Falta acta de nacimiento"
    materias_adeudadas = db.Column(db.Text, nullable=True)  # Nota manual; el cálculo automático llega con el Historial
    faltas_administrativas = db.Column(db.Text, nullable=True)  # Ej. "2 faltas por inasistencia a junta"

    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_validacion = db.Column(db.DateTime, nullable=True)  # Se llena cuando admin aprueba el pre-registro

    # Relaciones
    calificaciones = db.relationship(
        'Calificacion',
        backref='alumno',
        lazy=True,
        cascade='all, delete-orphan'
    )
    documentos = db.relationship(
        'DocumentoAlumno',
        backref='alumno',
        lazy=True,
        cascade='all, delete-orphan'
    )

    def __repr__(self):
        return f'<Alumno {self.matricula_id} - {self.nombre_completo}>'

    def edad(self) -> int:
        """Calcula la edad actual del alumno a partir de su fecha de nacimiento."""
        hoy = datetime.utcnow().date()
        anios = hoy.year - self.fecha_nacimiento.year
        # Ajusta si aún no ha cumplido años este año
        if (hoy.month, hoy.day) < (self.fecha_nacimiento.month, self.fecha_nacimiento.day):
            anios -= 1
        return anios

    def materia_pertenece_a_su_plan(self, materia: 'Materia') -> bool:
        """
        Método de apoyo para el 'Escudo' del Plan de Estudios.
        Se usará en el Módulo de Captura de Calificaciones (Paso 4) para
        rechazar cualquier intento de registrar una calificación de una
        materia que no corresponda al plan asignado a este alumno.
        """
        return materia.id_plan_fk == self.id_plan_fk

    def promedio_general(self) -> float:
        """Calcula el promedio general con las calificaciones ya capturadas."""
        if not self.calificaciones:
            return 0.0
        suma = sum(c.calificacion_final for c in self.calificaciones)
        return round(suma / len(self.calificaciones), 2)

    def saldo_total_adeudado(self):
        """
        Suma el saldo pendiente de TODOS los cargos no cancelados del
        alumno. Se define aquí (no en Cargo) porque necesita recorrer
        todos los cargos de este alumno específico.
        """
        return sum(
            (c.saldo_pendiente() for c in self.cargos if c.estatus != EstatusCargo.CANCELADO),
            Decimal('0.00')
        )

    def tiene_adeudo(self) -> bool:
        return self.saldo_total_adeudado() > 0


class TipoDocumento(enum.Enum):
    """Tipos de documentos digitalizados que se pueden anexar al expediente."""
    COMPROBANTE_DOMICILIO = 'Comprobante de Domicilio'
    INE = 'INE / Identificación Oficial'
    CURP_DOC = 'CURP (documento)'
    ACTA_NACIMIENTO = 'Acta de Nacimiento'
    CERTIFICADO_PREPA = 'Certificado de Preparatoria'
    FOTOGRAFIA = 'Fotografía'
    OTRO = 'Otro'


class DocumentoAlumno(db.Model):
    """
    Archivo digitalizado anexado al expediente de un Alumno
    (comprobante de domicilio, INE, CURP, etc.). El archivo físico se
    guarda en disco bajo UPLOAD_FOLDER/<matricula>/ y aquí solo se guarda
    la referencia/metadatos, para no inflar la base de datos con binarios.
    """
    __tablename__ = 'documentos_alumno'

    id = db.Column(db.Integer, primary_key=True)

    matricula_fk = db.Column(
        db.String(20),
        db.ForeignKey('alumnos.matricula_id'),
        nullable=False
    )

    tipo_documento = db.Column(db.Enum(TipoDocumento), nullable=False)
    nombre_archivo_original = db.Column(db.String(255), nullable=False)
    ruta_archivo = db.Column(db.String(500), nullable=False)  # Ruta relativa dentro de UPLOAD_FOLDER
    fecha_subida = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<DocumentoAlumno {self.matricula_fk} - {self.tipo_documento.value}>'


class Calificacion(db.Model):
    """
    Calificación final de un Alumno en una Materia, para un periodo escolar
    determinado. Se vincula SIEMPRE a la matrícula del alumno (no al grupo),
    cumpliendo la regla de negocio "Centrado en el Alumno".

    La restricción única (matricula_fk, id_materia_fk, periodo_escolar)
    evita capturas duplicadas accidentales de la misma acta.
    """
    __tablename__ = 'calificaciones'

    id = db.Column(db.Integer, primary_key=True)

    matricula_fk = db.Column(
        db.String(20),
        db.ForeignKey('alumnos.matricula_id'),
        nullable=False
    )
    id_materia_fk = db.Column(
        db.Integer,
        db.ForeignKey('materias.id'),
        nullable=False
    )

    calificacion_final = db.Column(db.Float, nullable=False)
    periodo_escolar = db.Column(db.String(20), nullable=False)  # Ej: "2025-B", "Ene-Abr 2025"

    fecha_captura = db.Column(db.DateTime, default=datetime.utcnow)
    capturado_por = db.Column(db.String(100), nullable=True)  # Usuario admin que capturó (auditoría)
    numero_acta = db.Column(db.String(50), nullable=True)  # Referencia al acta física firmada

    __table_args__ = (
        UniqueConstraint(
            'matricula_fk', 'id_materia_fk', 'periodo_escolar',
            name='uq_calificacion_alumno_materia_periodo'
        ),
        CheckConstraint(
            'calificacion_final >= 0 AND calificacion_final <= 10',
            name='ck_calificacion_rango_valido'
        ),
    )

    def __repr__(self):
        return f'<Calificacion {self.matricula_fk} / Materia {self.id_materia_fk}: {self.calificacion_final}>'


class HistorialEstatus(db.Model):
    """
    Auditoría de cada cambio de estatus de un alumno (ej. Pendiente -> Activo,
    Activo -> Egresado). Nunca se sobreescribe ni se borra: es un registro
    histórico de "quién cambió qué y cuándo".
    """
    __tablename__ = 'historial_estatus'

    id = db.Column(db.Integer, primary_key=True)

    matricula_fk = db.Column(
        db.String(20),
        db.ForeignKey('alumnos.matricula_id'),
        nullable=False
    )
    usuario_fk = db.Column(
        db.Integer,
        db.ForeignKey('usuarios.id'),
        nullable=True  # Nullable por si el usuario se llega a borrar en el futuro
    )

    estatus_anterior = db.Column(db.Enum(EstatusAlumno), nullable=True)  # Nulo en el primer registro
    estatus_nuevo = db.Column(db.Enum(EstatusAlumno), nullable=False)
    comentario = db.Column(db.String(255), nullable=True)  # Ej. "Egresado forzado sin todas las materias"
    fecha = db.Column(db.DateTime, default=datetime.utcnow)

    alumno = db.relationship('Alumno', backref=db.backref('historial_estatus', lazy=True, order_by='HistorialEstatus.fecha.desc()'))
    usuario = db.relationship('Usuario')

    def __repr__(self):
        return f'<HistorialEstatus {self.matricula_fk}: {self.estatus_anterior} -> {self.estatus_nuevo}>'


# ---------------------------------------------------------------------------
# SISTEMA DE COBROS (colegiaturas, inscripción, etc.)
# Alcance deliberadamente acotado: cargos y pagos por alumno (NO es un
# sistema contable de activos/pasivos de la universidad — eso, si algún
# día se necesita, sería un módulo aparte y mucho más grande).
# ---------------------------------------------------------------------------

class EstatusCargo(enum.Enum):
    PENDIENTE = 'Pendiente'
    PARCIAL = 'Pago Parcial'
    PAGADO = 'Pagado'
    CANCELADO = 'Cancelado'


class MetodoPago(enum.Enum):
    EFECTIVO = 'Efectivo'
    TRANSFERENCIA = 'Transferencia Bancaria'
    TARJETA = 'Tarjeta'
    DEPOSITO = 'Depósito Bancario'
    OTRO = 'Otro'


class TipoRecargo(enum.Enum):
    """Cada universidad calcula sus recargos distinto — por eso es configurable."""
    MONTO_FIJO = 'Monto fijo (una sola vez)'
    PORCENTAJE = 'Porcentaje del adeudo'
    POR_DIA = 'Monto fijo por cada día de atraso'
    PORCENTAJE_MENSUAL = 'Porcentaje acumulativo por cada mes de atraso'


class ConceptoCobro(db.Model):
    """
    Catálogo de conceptos de cobro (Colegiatura, Inscripción, Servicio
    Social, Uniformes, Recargo, etc.). Se administra desde la interfaz —
    NUNCA se escribe como texto libre al capturar un cargo, para que los
    reportes por concepto sean siempre consistentes.
    """
    __tablename__ = 'conceptos_cobro'

    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True, nullable=False)
    activo = db.Column(db.Boolean, default=True, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<ConceptoCobro {self.nombre}>'


class ConfiguracionCobros(db.Model):
    """
    Configuración GLOBAL de recargos por atraso. Es una tabla de una sola
    fila (patrón "singleton"): siempre se usa/edita el único registro que
    exista, vía ConfiguracionCobros.obtener(). Así el recargo es
    "auto-ajustable" por institución sin tocar código — Dirección lo
    cambia desde una pantalla y aplica de inmediato a todos los cargos
    vencidos de ahí en adelante.
    """
    __tablename__ = 'configuracion_cobros'

    id = db.Column(db.Integer, primary_key=True)
    tipo_recargo = db.Column(db.Enum(TipoRecargo), nullable=False, default=TipoRecargo.MONTO_FIJO)
    valor_recargo = db.Column(db.Numeric(10, 2), nullable=False, default=Decimal('0.00'))
    dias_gracia = db.Column(db.Integer, nullable=False, default=0)  # Días después del vencimiento antes de recargar

    @staticmethod
    def obtener():
        """Devuelve la única configuración existente, creándola con valores neutros si no existe."""
        config = ConfiguracionCobros.query.first()
        if not config:
            config = ConfiguracionCobros(
                tipo_recargo=TipoRecargo.MONTO_FIJO,
                valor_recargo=Decimal('0.00'),
                dias_gracia=0,
            )
            db.session.add(config)
            db.session.commit()
        return config


class Cargo(db.Model):
    """
    Un cobro pendiente para un alumno (colegiatura de un mes, inscripción,
    reinscripción, materiales, etc.). Puede pagarse en una sola exhibición
    o en varios pagos parciales — el saldo y el estatus se calculan
    siempre a partir de la suma de sus Pago asociados, nunca se guardan
    "a mano", para que nunca queden desincronizados.
    """
    __tablename__ = 'cargos'

    id = db.Column(db.Integer, primary_key=True)
    matricula_fk = db.Column(db.String(20), db.ForeignKey('alumnos.matricula_id'), nullable=False)

    concepto_cobro_fk = db.Column(db.Integer, db.ForeignKey('conceptos_cobro.id'), nullable=True)
    concepto = db.Column(db.String(150), nullable=False)  # Denormalizado: nombre del concepto al momento de crear el cargo
    monto = db.Column(db.Numeric(10, 2), nullable=False)
    recargo_aplicado = db.Column(db.Numeric(10, 2), nullable=False, default=Decimal('0.00'))
    periodo_escolar = db.Column(db.String(20), nullable=True)
    fecha_vencimiento = db.Column(db.Date, nullable=True)
    fecha_generacion = db.Column(db.DateTime, default=datetime.utcnow)
    estatus = db.Column(db.Enum(EstatusCargo), default=EstatusCargo.PENDIENTE, nullable=False)
    comentario = db.Column(db.String(255), nullable=True)  # Ej. motivo de cancelación

    generado_por_fk = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)

    alumno = db.relationship(
        'Alumno',
        backref=db.backref('cargos', lazy=True, order_by='Cargo.fecha_generacion.desc()')
    )
    generado_por = db.relationship('Usuario')
    concepto_cobro = db.relationship('ConceptoCobro')

    def total_pagado(self):
        return sum((p.monto_pagado for p in self.pagos), Decimal('0.00'))

    def saldo_pendiente(self):
        return (self.monto + self.recargo_aplicado) - self.total_pagado()

    def actualizar_estatus(self):
        """Recalcula el estatus a partir de los pagos reales. Nunca se sobreescribe a mano."""
        if self.estatus == EstatusCargo.CANCELADO:
            return
        saldo = self.saldo_pendiente()
        if saldo <= 0:
            self.estatus = EstatusCargo.PAGADO
        elif self.total_pagado() > 0:
            self.estatus = EstatusCargo.PARCIAL
        else:
            self.estatus = EstatusCargo.PENDIENTE

    def esta_vencido(self):
        """Calculado en tiempo real (no se guarda) para no depender de un job en segundo plano."""
        return (
            self.estatus not in (EstatusCargo.PAGADO, EstatusCargo.CANCELADO)
            and self.fecha_vencimiento is not None
            and self.fecha_vencimiento < datetime.utcnow().date()
        )

    def actualizar_recargo_si_vencido(self):
        """
        Recalcula el recargo con la configuración VIGENTE (auto-ajustable:
        si Dirección cambia la fórmula, aplica de inmediato a partir de
        aquí). El recargo nunca DISMINUYE aunque la config cambie a la
        baja — solo puede crecer o quedarse igual, para no borrar
        recargos que ya se hicieron oficiales en algún reporte previo.
        Se llama cada vez que se listan los cargos de un alumno (sin
        necesidad de un cron job en segundo plano).
        """
        if self.estatus in (EstatusCargo.PAGADO, EstatusCargo.CANCELADO):
            return
        if not self.fecha_vencimiento:
            return

        config = ConfiguracionCobros.obtener()
        hoy = datetime.utcnow().date()
        dias_de_atraso = (hoy - self.fecha_vencimiento).days - config.dias_gracia

        if config.tipo_recargo == TipoRecargo.MONTO_FIJO:
            recargo_calculado = config.valor_recargo
        elif config.tipo_recargo == TipoRecargo.PORCENTAJE:
            recargo_calculado = (self.monto * config.valor_recargo / Decimal('100')).quantize(Decimal('0.01'))
        elif config.tipo_recargo == TipoRecargo.POR_DIA:
            recargo_calculado = (config.valor_recargo * dias_de_atraso).quantize(Decimal('0.01'))
        elif config.tipo_recargo == TipoRecargo.PORCENTAJE_MENSUAL:
            # Ej.: mensualidad $2,000, valor_recargo=10 -> se suman $200 por
            # cada mes COMPLETO de atraso (mes 1: $200, mes 2: $400, etc.).
            # Aproximamos "mes" como bloques de 30 días de atraso.
            meses_de_atraso = -(-dias_de_atraso // 30)  # redondeo hacia arriba, mínimo 1
            recargo_calculado = (
                self.monto * config.valor_recargo / Decimal('100') * meses_de_atraso
            ).quantize(Decimal('0.01'))
        else:
            recargo_calculado = Decimal('0.00')

        if recargo_calculado > self.recargo_aplicado:
            self.recargo_aplicado = recargo_calculado

    def __repr__(self):
        return f'<Cargo {self.matricula_fk}: {self.concepto} - ${self.monto}>'


class Pago(db.Model):
    """Un pago (total o parcial) aplicado a un Cargo específico."""
    __tablename__ = 'pagos'

    id = db.Column(db.Integer, primary_key=True)
    cargo_fk = db.Column(db.Integer, db.ForeignKey('cargos.id'), nullable=False)

    folio = db.Column(db.String(30), unique=True, nullable=True)  # Ej. "PAGO-2026-000042"
    monto_pagado = db.Column(db.Numeric(10, 2), nullable=False)
    fecha_pago = db.Column(db.DateTime, default=datetime.utcnow)
    metodo_pago = db.Column(db.Enum(MetodoPago), nullable=False, default=MetodoPago.EFECTIVO)
    referencia = db.Column(db.String(100), nullable=True)  # Folio/número de referencia bancaria
    comentario = db.Column(db.String(255), nullable=True)

    capturado_por_fk = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)

    cargo = db.relationship(
        'Cargo',
        backref=db.backref('pagos', lazy=True, order_by='Pago.fecha_pago.desc()')
    )
    capturado_por = db.relationship('Usuario')

    def __repr__(self):
        return f'<Pago {self.cargo_fk}: ${self.monto_pagado}>'


# ---------------------------------------------------------------------------
# APPLICATION FACTORY
# ---------------------------------------------------------------------------

def create_app(config_name='development'):
    app = Flask(__name__, instance_relative_config=True)
    app.config.from_object(config_by_name[config_name])

    # SECURITY-NOTE: aquí SÍ se ejecuta esta validación (a diferencia de un
    # __init__ en config.py, que Flask jamás llamaría porque from_object()
    # recibe la CLASE, no una instancia). Si en producción falta SECRET_KEY,
    # preferimos que la app truene al arrancar a que arranque en silencio
    # con una clave insegura y públicamente conocida.
    if config_name == 'production' and not app.config.get('SECRET_KEY'):
        raise RuntimeError(
            'SECRET_KEY no está definida en el entorno de producción. '
            'Genera una clave larga y aleatoria (ej. con `python -c '
            '"import secrets; print(secrets.token_hex(32))"`) y agrégala '
            'a tu archivo .env antes de arrancar la aplicación.'
        )

    if config_name == 'production':
        # El VPS sirve la app detrás de Nginx (proxy inverso). Sin esto,
        # Flask-Limiter (get_remote_address) vería SIEMPRE la IP de Nginx
        # en vez de la IP real del navegador, y el límite de intentos de
        # login se compartiría entre TODOS los usuarios en vez de aplicarse
        # por IP real. x_for=1 confía en un solo salto de proxy (ajusta si
        # tu VPS tiene más de un proxy intermedio, ej. Cloudflare + Nginx).
        app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

    db.init_app(app)
    migrate.init_app(app, db)

    login_manager.init_app(app)
    login_manager.login_view = 'login'
    login_manager.login_message = 'Debes iniciar sesión para acceder a esta página.'
    login_manager.login_message_category = 'warning'

    csrf.init_app(app)
    limiter.init_app(app)
    mail.init_app(app)

    # Asegura que exista la carpeta física donde se guardan los documentos
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    # Aquí se registrarán los Blueprints en pasos posteriores:
    # from routes.registro import registro_bp
    # from routes.admin import admin_bp
    # app.register_blueprint(registro_bp)
    # app.register_blueprint(admin_bp)

    return app


app = create_app(os.environ.get('FLASK_ENV', 'development'))


@app.errorhandler(429)
def limite_intentos_excedido(error):
    """
    Se dispara cuando Flask-Limiter bloquea una IP por exceder el límite de
    intentos de login. En vez del 429 genérico de Werkzeug, mostramos un
    mensaje claro y regresamos a la pantalla de login.
    """
    flash('Demasiados intentos de inicio de sesión. Por seguridad, espera un minuto e inténtalo de nuevo.', 'danger')
    return redirect(url_for('login'))


@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))


def es_url_segura(destino: str) -> bool:
    """
    Evita un 'Open Redirect': sin esta validación, alguien podría mandar un
    link tipo /login?next=https://sitio-falso.com y, tras iniciar sesión
    correctamente en el sitio REAL, el usuario terminaría redirigido a un
    sitio externo (útil para phishing dirigido al personal). Solo se
    permite continuar si 'next' es una ruta relativa de este mismo sitio
    (sin esquema http/https ni host propios).
    """
    if not destino:
        return False
    partes = urlparse(destino)
    return not partes.scheme and not partes.netloc


def rol_requerido(*roles_permitidos):
    """
    Decorador para restringir una ruta a ciertos roles (ej. solo DIRECTIVO).
    Siempre exige login primero (@login_required incluido). Uso:

        @app.route('/algo-solo-de-dirección')
        @rol_requerido('DIRECTIVO')
        def algo():
            ...
    """
    def decorador(func):
        @wraps(func)
        @login_required
        def envoltura(*args, **kwargs):
            if current_user.rol.name not in roles_permitidos:
                flash('No tienes permisos para realizar esta acción.', 'danger')
                return redirect(url_for('index'))
            return func(*args, **kwargs)
        return envoltura
    return decorador


# ---------------------------------------------------------------------------
# AUTENTICACIÓN
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit('5 per minute', methods=['POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        recordar = request.form.get('recordar') == 'on'

        usuario = Usuario.query.filter_by(username=username).first()

        # Mensaje genérico a propósito: no revelamos si falló el usuario
        # o la contraseña, para no facilitar enumeración de cuentas.
        if usuario and usuario.activo and usuario.check_password(password):
            session.permanent = True  # Activa PERMANENT_SESSION_LIFETIME (expira tras 8h de inactividad)
            login_user(usuario, remember=recordar)
            usuario.ultimo_acceso = datetime.utcnow()
            db.session.commit()

            flash(f'Bienvenido, {usuario.nombre_completo}.', 'success')
            siguiente = request.args.get('next')
            destino = siguiente if es_url_segura(siguiente) else url_for('index')
            return redirect(destino)

        flash('Usuario o contraseña incorrectos.', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sesión cerrada correctamente.', 'success')
    return redirect(url_for('login'))


@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    """Cada usuario (Directivo o Administrativo) edita SU propia cuenta aquí."""
    if request.method == 'POST':
        nombre_completo = request.form.get('nombre_completo', '').strip()
        if nombre_completo:
            current_user.nombre_completo = nombre_completo

        password_actual = request.form.get('password_actual', '')
        password_nueva = request.form.get('password_nueva', '')
        password_confirmar = request.form.get('password_confirmar', '')

        if password_nueva or password_confirmar or password_actual:
            if not current_user.check_password(password_actual):
                flash('Tu contraseña actual no es correcta; no se cambió nada.', 'danger')
                return redirect(url_for('perfil'))
            if len(password_nueva) < 8:
                flash('La nueva contraseña debe tener al menos 8 caracteres.', 'danger')
                return redirect(url_for('perfil'))
            if password_nueva != password_confirmar:
                flash('La confirmación no coincide con la nueva contraseña.', 'danger')
                return redirect(url_for('perfil'))
            current_user.set_password(password_nueva)
            flash('Contraseña actualizada correctamente.', 'success')

        db.session.commit()
        flash('Perfil actualizado.', 'success')
        return redirect(url_for('perfil'))

    return render_template('perfil.html')


# ---------------------------------------------------------------------------
# GESTIÓN DE ADMINISTRATIVOS (solo DIRECTIVO)
# ---------------------------------------------------------------------------

@app.route('/usuarios')
@rol_requerido('DIRECTIVO')
def usuarios():
    lista = Usuario.query.order_by(Usuario.rol.asc(), Usuario.nombre_completo.asc()).all()
    return render_template('usuarios.html', usuarios=lista)


@app.route('/usuarios/nuevo', methods=['GET', 'POST'])
@rol_requerido('DIRECTIVO')
def nuevo_usuario():
    if request.method == 'POST':
        nombre_completo = request.form.get('nombre_completo', '').strip()
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        confirmar = request.form.get('confirmar_password', '')
        rol_raw = request.form.get('rol', 'ADMINISTRATIVO')

        errores = []
        if len(nombre_completo) < 3:
            errores.append('Ingresa el nombre completo del usuario.')
        if len(username) < 4 or not re.match(r'^[a-z0-9_.]+$', username):
            errores.append('El usuario debe tener al menos 4 caracteres (letras, números, "." o "_").')
        elif Usuario.query.filter_by(username=username).first():
            errores.append(f'El usuario "{username}" ya existe.')
        if len(password) < 8:
            errores.append('La contraseña debe tener al menos 8 caracteres.')
        elif password != confirmar:
            errores.append('Las contraseñas no coinciden.')
        if rol_raw not in RolUsuario.__members__:
            errores.append('Selecciona un rol válido.')

        if errores:
            for error in errores:
                flash(error, 'danger')
            return render_template('nuevo_usuario.html'), 400

        nuevo = Usuario(
            nombre_completo=nombre_completo,
            username=username,
            rol=RolUsuario[rol_raw],
            activo=True
        )
        nuevo.set_password(password)

        try:
            db.session.add(nuevo)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash('Ocurrió un error al crear el usuario. Intenta de nuevo.', 'danger')
            return render_template('nuevo_usuario.html'), 400

        flash(f'Usuario "{username}" ({nuevo.rol.value}) creado correctamente.', 'success')
        return redirect(url_for('usuarios'))

    return render_template('nuevo_usuario.html')


@app.route('/usuarios/<int:user_id>/toggle', methods=['POST'])
@rol_requerido('DIRECTIVO')
def toggle_usuario(user_id):
    """Activa/desactiva una cuenta sin borrarla (mejor que eliminarla: conserva auditoría)."""
    usuario = Usuario.query.get_or_404(user_id)

    if usuario.id == current_user.id:
        flash('No puedes desactivar tu propia cuenta.', 'danger')
        return redirect(url_for('usuarios'))

    usuario.activo = not usuario.activo
    db.session.commit()

    estado = 'activada' if usuario.activo else 'desactivada'
    flash(f'La cuenta de "{usuario.username}" fue {estado}.', 'success')
    return redirect(url_for('usuarios'))


# ---------------------------------------------------------------------------
# MÓDULO DE BÚSQUEDA UNIVERSAL
# Pantalla principal de uso exclusivo de Control Escolar.
# ---------------------------------------------------------------------------

def _matriculas_con_adeudo():
    """
    Devuelve el conjunto de matrículas con saldo pendiente > 0, calculado
    con UNA agregación en SQL (subquery de pagos por cargo) en vez de
    cargar cada Alumno y cada Cargo/Pago en Python.

    PERFORMANCE-NOTE: la versión anterior hacía Alumno.query.join(Cargo)...all()
    y luego, para cada alumno, tiene_adeudo() -> saldo_total_adeudado(),
    que a su vez recorre alumno.cargos y, por cada cargo, cargo.pagos
    (ambos lazy='select'). Con pocos alumnos no se nota, pero es un
    patrón N+1 clásico: con miles de alumnos con adeudo, cada carga del
    dashboard dispara cientos/miles de consultas adicionales. Esta versión
    hace el cálculo en 1 sola consulta agregada, sin importar cuántos
    alumnos existan.
    """
    pagos_por_cargo = (
        db.session.query(
            Pago.cargo_fk,
            func.coalesce(func.sum(Pago.monto_pagado), 0).label('total_pagado')
        )
        .group_by(Pago.cargo_fk)
        .subquery()
    )

    saldo_expr = (
        Cargo.monto + Cargo.recargo_aplicado
        - func.coalesce(pagos_por_cargo.c.total_pagado, 0)
    )

    filas = (
        db.session.query(Cargo.matricula_fk, func.sum(saldo_expr).label('saldo'))
        .outerjoin(pagos_por_cargo, pagos_por_cargo.c.cargo_fk == Cargo.id)
        .filter(Cargo.estatus != EstatusCargo.CANCELADO)
        .group_by(Cargo.matricula_fk)
        .having(func.sum(saldo_expr) > 0)
        .all()
    )
    return {matricula for matricula, _saldo in filas}


def calcular_estadisticas_alumnos():
    """Números clave para el panel del buscador (pantalla de inicio)."""
    con_adeudo = len(_matriculas_con_adeudo())

    return {
        'total': Alumno.query.count(),
        'pendientes': Alumno.query.filter_by(estatus=EstatusAlumno.PENDIENTE).count(),
        'activos': Alumno.query.filter_by(estatus=EstatusAlumno.ACTIVO).count(),
        'documentacion_incompleta': Alumno.query.filter(Alumno.documentacion_pendiente.isnot(None)).count(),
        'faltas': Alumno.query.filter(Alumno.faltas_administrativas.isnot(None)).count(),
        'con_adeudo': con_adeudo,
    }


@app.route('/')
@login_required
def index():
    """
    Pantalla de inicio: dashboard con números clave + filtros rápidos,
    y el buscador universal. Si viene ?filtro=algo en la URL, muestra
    esa lista filtrada en vez del dashboard vacío.
    """
    estadisticas = calcular_estadisticas_alumnos()

    filtros_disponibles = {
        'pendientes': ('Alumnos Pendientes de Validación', Alumno.estatus == EstatusAlumno.PENDIENTE),
        'activos': ('Alumnos Activos', Alumno.estatus == EstatusAlumno.ACTIVO),
        'documentacion': ('Alumnos con Documentación Pendiente', Alumno.documentacion_pendiente.isnot(None)),
        'faltas': ('Alumnos con Faltas Administrativas', Alumno.faltas_administrativas.isnot(None)),
        'con_adeudo': ('Alumnos con Adeudo Económico', None),
        'recientes': ('Últimos 10 Alumnos Registrados', None),
    }

    filtro = request.args.get('filtro')
    resultados = None
    titulo_filtro = None

    if filtro in filtros_disponibles:
        titulo_filtro, condicion = filtros_disponibles[filtro]
        if filtro == 'recientes':
            resultados = Alumno.query.order_by(Alumno.fecha_registro.desc()).limit(10).all()
        elif filtro == 'con_adeudo':
            matriculas = _matriculas_con_adeudo()
            resultados = (
                Alumno.query
                .filter(Alumno.matricula_id.in_(matriculas))
                .order_by(Alumno.nombre_completo.asc())
                .all()
            ) if matriculas else []
        else:
            resultados = Alumno.query.filter(condicion).order_by(Alumno.nombre_completo.asc()).all()

    return render_template(
        'buscador.html',
        estadisticas=estadisticas,
        resultados=resultados,
        titulo_filtro=titulo_filtro
    )


@app.route('/buscar', methods=['POST'])
@login_required
def buscar():
    """
    Búsqueda flexible tipo 'banco':
      - Coincidencia EXACTA por matricula_id.
      - O coincidencia PARCIAL (case-insensitive) por nombre_completo.
    """
    termino = request.form.get('termino', '').strip()

    if not termino:
        flash('Ingresa una matrícula o un nombre para buscar.', 'warning')
        return redirect(url_for('index'))

    resultados = Alumno.query.filter(
        or_(
            Alumno.matricula_id == termino,
            Alumno.nombre_completo.ilike(f'%{termino}%')
        )
    ).order_by(Alumno.nombre_completo.asc()).all()

    if not resultados:
        flash(f'No se encontraron alumnos que coincidan con "{termino}".', 'danger')
        return redirect(url_for('index'))

    return render_template('buscador.html', resultados=resultados, termino=termino)


# ---------------------------------------------------------------------------
# MÓDULO DE AUTO-REGISTRO PÚBLICO
# ---------------------------------------------------------------------------

def generar_matricula(plan: 'PlanEstudio') -> str:
    """
    Genera la matrícula siguiente para un plan dado, con el formato:
        <CLAVE_CARRERA><AÑO_ACTUAL>-<CONSECUTIVO 5 dígitos>
    Ej: LEN2026-00001, LEN2026-00002, ...
    El consecutivo se calcula buscando la última matrícula ya usada
    con ese mismo prefijo (carrera + año), NO el total de alumnos,
    para que nunca se reutilice un número aunque se den de baja alumnos.

    NOTA sobre concurrencia: with_for_update() bloquea la fila leída para
    que otra transacción no pueda leer el mismo "último folio" hasta que
    esta termine — esto reduce la condición de carrera en PostgreSQL
    (producción). En SQLite (desarrollo) no hay bloqueo por fila, así que
    la cláusula se ignora silenciosamente sin causar error; por eso la
    protección real y definitiva contra colisiones es la función
    crear_alumno_generando_matricula() de abajo, que reintenta si de
    todos modos ocurre un choque (ej. dos registros "primeros" del mismo
    prefijo, exactamente al mismo tiempo, donde no hay fila que bloquear).
    """
    anio_actual = datetime.utcnow().year
    prefijo = f"{plan.clave_carrera}{anio_actual}-"

    consulta = (
        Alumno.query
        .filter(Alumno.matricula_id.like(f"{prefijo}%"))
        .order_by(Alumno.matricula_id.desc())
    )

    # with_for_update() (bloqueo de fila) solo tiene efecto real en motores
    # que lo soportan, como PostgreSQL (producción). SQLite (desarrollo)
    # no tiene bloqueo por fila — en vez de asumir que SQLAlchemy lo ignora
    # solo, lo excluimos explícitamente aquí para no depender de un
    # comportamiento de dialecto sin poder verificarlo en este entorno.
    if db.engine.dialect.name != 'sqlite':
        consulta = consulta.with_for_update()

    ultimo = consulta.first()

    if ultimo:
        ultimo_num = int(ultimo.matricula_id.split('-')[-1])
        siguiente = ultimo_num + 1
    else:
        siguiente = 1

    return f"{prefijo}{siguiente:05d}"


def crear_alumno_generando_matricula(plan: 'PlanEstudio', intentos_maximos: int = 3, **datos_alumno):
    """
    Crea y guarda un Alumno generándole matrícula automáticamente, con
    reintentos ante una colisión de matrícula por condición de carrera
    (dos registros al mismo tiempo). Hace su PROPIO commit (por diseño:
    así, si se usa dentro de un bucle de carga masiva, una fila que falla
    nunca arrastra consigo a las filas anteriores que ya se guardaron bien
    — cada una queda comprometida en la base de datos de forma individual).

    Devuelve (alumno, None) si tuvo éxito, o (None, mensaje_error) si
    fallaron todos los intentos. NO valida CURP duplicada ni nada del
    resto de las reglas de negocio — eso debe hacerse ANTES de llamar
    aquí; esta función solo protege la generación de matrícula.
    """
    for _intento in range(intentos_maximos):
        matricula = generar_matricula(plan)
        nuevo_alumno = Alumno(matricula_id=matricula, id_plan_fk=plan.id, **datos_alumno)

        try:
            db.session.add(nuevo_alumno)
            db.session.commit()
            return nuevo_alumno, None
        except IntegrityError:
            db.session.rollback()
            continue  # Probable colisión de matrícula: se reintenta con el siguiente consecutivo

    return None, 'no se pudo generar una matrícula única tras varios intentos; intenta de nuevo'


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    planes = PlanEstudio.query.filter_by(activo=True).order_by(PlanEstudio.nombre.asc()).all()

    if request.method == 'GET':
        return render_template('registro.html', planes=planes)

    # --- POST: procesar el formulario ---
    nombre_completo = request.form.get('nombre_completo', '').strip()
    curp = request.form.get('curp', '').strip().upper()
    fecha_nacimiento_raw = request.form.get('fecha_nacimiento', '')
    fecha_certificado_raw = request.form.get('fecha_certificado_prepa', '')
    id_plan_raw = request.form.get('id_plan_fk', '')
    correo = request.form.get('correo', '').strip() or None
    telefono = request.form.get('telefono', '').strip() or None
    telefono_movil = request.form.get('telefono_movil', '').strip() or None

    sexo = request.form.get('sexo', '').strip() or None
    numero_identificacion = request.form.get('numero_identificacion', '').strip() or None
    estado_civil = request.form.get('estado_civil', '').strip() or None
    nacionalidad = request.form.get('nacionalidad', '').strip() or 'Mexicana'
    tipo_sangre = request.form.get('tipo_sangre', '').strip() or None

    domicilio_calle_numero = request.form.get('domicilio_calle_numero', '').strip() or None
    domicilio_ciudad = request.form.get('domicilio_ciudad', '').strip() or None
    domicilio_cp = request.form.get('domicilio_cp', '').strip() or None
    domicilio_estado = request.form.get('domicilio_estado', '').strip() or None

    contacto_emergencia_nombre = request.form.get('contacto_emergencia_nombre', '').strip() or None
    contacto_emergencia_telefono = request.form.get('contacto_emergencia_telefono', '').strip() or None
    contacto_emergencia_parentesco = request.form.get('contacto_emergencia_parentesco', '').strip() or None

    como_se_entero = request.form.get('como_se_entero', '').strip() or None

    turno_raw = request.form.get('turno', '')
    modalidad_raw = request.form.get('modalidad', '')

    errores = []

    if len(nombre_completo) < 5:
        errores.append('Ingresa tu nombre completo correctamente.')

    if not re.match(r'^[A-Z0-9]{18}$', curp):
        errores.append('La CURP debe tener exactamente 18 caracteres alfanuméricos.')
    elif Alumno.query.filter_by(curp=curp).first():
        errores.append(f'Ya existe un alumno registrado con la CURP "{curp}".')

    if not sexo:
        errores.append('Selecciona tu sexo.')

    if not domicilio_calle_numero or not domicilio_ciudad or not domicilio_cp or not domicilio_estado:
        errores.append('Completa todos los campos de tu domicilio.')

    if not contacto_emergencia_nombre or not contacto_emergencia_telefono:
        errores.append('Indica el nombre y teléfono de tu contacto de emergencia.')

    turno = None
    if turno_raw not in TurnoAlumno.__members__:
        errores.append('Selecciona un turno válido.')
    else:
        turno = TurnoAlumno[turno_raw]

    modalidad = None
    if modalidad_raw not in ModalidadEstudio.__members__:
        errores.append('Selecciona una modalidad válida.')
    else:
        modalidad = ModalidadEstudio[modalidad_raw]

    fecha_nacimiento = None
    try:
        fecha_nacimiento = datetime.strptime(fecha_nacimiento_raw, '%Y-%m-%d').date()
    except ValueError:
        errores.append('La fecha de nacimiento no es válida.')

    fecha_certificado_prepa = None
    try:
        fecha_certificado_prepa = datetime.strptime(fecha_certificado_raw, '%Y-%m-%d').date()
    except ValueError:
        errores.append('La fecha del certificado de preparatoria no es válida.')

    plan = None
    if not id_plan_raw:
        errores.append('Debes seleccionar tu carrera / plan de estudios.')
    else:
        try:
            plan = PlanEstudio.query.get(int(id_plan_raw))
        except (ValueError, TypeError):
            plan = None
        if not plan or not plan.activo:
            errores.append('El plan de estudios seleccionado no es válido.')

    if errores:
        for error in errores:
            flash(error, 'danger')
        # Reenviamos el formulario con los planes para no perder el <select>
        return render_template('registro.html', planes=planes), 400

    alumno, error_creacion = crear_alumno_generando_matricula(
        plan,
        nombre_completo=nombre_completo,
        curp=curp,
        fecha_nacimiento=fecha_nacimiento,
        fecha_certificado_prepa=fecha_certificado_prepa,
        estatus=EstatusAlumno.PENDIENTE,
        correo=correo,
        telefono=telefono,
        telefono_movil=telefono_movil,
        sexo=sexo,
        numero_identificacion=numero_identificacion,
        estado_civil=estado_civil,
        nacionalidad=nacionalidad,
        tipo_sangre=tipo_sangre,
        domicilio_calle_numero=domicilio_calle_numero,
        domicilio_ciudad=domicilio_ciudad,
        domicilio_cp=domicilio_cp,
        domicilio_estado=domicilio_estado,
        contacto_emergencia_nombre=contacto_emergencia_nombre,
        contacto_emergencia_telefono=contacto_emergencia_telefono,
        contacto_emergencia_parentesco=contacto_emergencia_parentesco,
        como_se_entero=como_se_entero,
        turno=turno,
        modalidad=modalidad,
    )

    if error_creacion:
        flash('Ocurrió un error al guardar tu registro. Verifica tus datos e intenta de nuevo.', 'danger')
        return render_template('registro.html', planes=planes), 400

    # SECURITY-NOTE: esta es la ÚNICA vista PÚBLICA sin login del sistema, así
    # que es la de mayor exposición. Antes la plantilla usaba {{ message|safe }}
    # para poder mostrar <strong>{matricula}</strong> en negritas -- pero eso
    # dejaba la puerta abierta a que un flash() futuro con datos de usuario sin
    # escapar se convirtiera en XSS reflejado. Ahora se arma explícitamente con
    # Markup() + escape(): el HTML fijo (las etiquetas <strong>) se conserva,
    # pero cualquier dato variable (la matrícula) SIEMPRE pasa por escape().
    flash(
        Markup(
            f'¡Registro exitoso! Tu matrícula es <strong>{escape(alumno.matricula_id)}</strong>. '
            'Tu solicitud quedó en estatus "Pendiente de Validación" y será revisada por Control Escolar.'
        ),
        'success'
    )
    return redirect(url_for('registro'))


# ---------------------------------------------------------------------------
# MÓDULO DE CARGA MASIVA DE ALUMNOS (Excel)
# Solo Directivo. Pensado para migrar alumnos YA inscritos con expediente
# físico, sin tener que registrarlos uno por uno desde el formulario web.
# La matrícula se genera automáticamente igual que en el registro público
# (reutiliza generar_matricula), y cada fila pasa por las MISMAS reglas
# de validación que el registro público (CURP, fechas, plan válido).
# ---------------------------------------------------------------------------

# (columna, es_obligatoria, descripción para la hoja de ayuda de la plantilla)
COLUMNAS_IMPORTACION_ALUMNOS = [
    ('nombre_completo', True, 'Nombre completo del alumno'),
    ('curp', True, 'CURP, 18 caracteres'),
    ('fecha_nacimiento', True, 'Formato AAAA-MM-DD, ej. 2005-03-21'),
    ('fecha_certificado_prepa', True, 'Formato AAAA-MM-DD'),
    ('clave_carrera', True, 'Clave del plan de estudios (ver hoja "Guía" para las claves válidas)'),
    ('sexo', True, 'Femenino / Masculino / Otro'),
    ('estatus', False, 'PENDIENTE / ACTIVO / BAJA_TEMPORAL / BAJA_DEFINITIVA / EGRESADO (vacío = ACTIVO)'),
    ('cuatrimestre_actual', False, 'Número de cuatrimestre en el que va (vacío = 1)'),
    ('correo', False, ''),
    ('telefono', False, 'Teléfono fijo'),
    ('telefono_movil', False, ''),
    ('domicilio_calle_numero', False, ''),
    ('domicilio_ciudad', False, ''),
    ('domicilio_cp', False, ''),
    ('domicilio_estado', False, ''),
    ('numero_identificacion', False, 'INE u otra identificación oficial'),
    ('estado_civil', False, ''),
    ('nacionalidad', False, 'Vacío = "Mexicana"'),
    ('tipo_sangre', False, 'Ej. O+, A-'),
    ('contacto_emergencia_nombre', False, ''),
    ('contacto_emergencia_telefono', False, ''),
    ('contacto_emergencia_parentesco', False, 'Ej. Madre, Hermana'),
    ('turno', False, 'MATUTINO / VESPERTINO / MIXTO'),
    ('modalidad', False, 'ESCOLARIZADO / SEMIESCOLARIZADO / DISTANCIA'),
    ('grupo_actual', False, ''),
]


@app.route('/alumnos/importar/plantilla')
@rol_requerido('DIRECTIVO')
def plantilla_importacion():
    """Genera y descarga el archivo .xlsx en blanco con las columnas esperadas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Alumnos'

    fuente_encabezado = Font(bold=True, color='FFFFFF')
    relleno_obligatorio = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    relleno_opcional = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')

    for col_idx, (nombre_col, requerido, _desc) in enumerate(COLUMNAS_IMPORTACION_ALUMNOS, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre_col)
        celda.font = fuente_encabezado
        celda.fill = relleno_obligatorio if requerido else relleno_opcional
        ws.column_dimensions[get_column_letter(col_idx)].width = 26

    ws.freeze_panes = 'A2'

    # --- Hoja de ayuda: qué significa cada columna + claves de carrera válidas ---
    ws_guia = wb.create_sheet('Guía')
    ws_guia.append(['Columna', 'Obligatoria', 'Descripción'])
    for celda in ws_guia[1]:
        celda.font = Font(bold=True)

    for nombre_col, requerido, desc in COLUMNAS_IMPORTACION_ALUMNOS:
        ws_guia.append([nombre_col, 'Sí' if requerido else 'No', desc])

    ws_guia.column_dimensions['A'].width = 28
    ws_guia.column_dimensions['B'].width = 12
    ws_guia.column_dimensions['C'].width = 60

    ws_guia.append([])
    fila_titulo_planes = ws_guia.max_row + 1
    ws_guia.cell(row=fila_titulo_planes, column=1, value='Claves de carrera disponibles:').font = Font(bold=True)

    for plan in PlanEstudio.query.filter_by(activo=True).order_by(PlanEstudio.clave_carrera.asc()).all():
        ws_guia.append([plan.clave_carrera, plan.nombre])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='plantilla_carga_masiva_alumnos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/alumnos/importar', methods=['GET', 'POST'])
@rol_requerido('DIRECTIVO')
def importar_alumnos():
    if request.method == 'GET':
        return render_template('importar_alumnos.html')

    archivo = request.files.get('archivo_excel')
    if not archivo or archivo.filename == '':
        flash('Selecciona un archivo Excel (.xlsx) para importar.', 'danger')
        return redirect(url_for('importar_alumnos'))

    if not archivo.filename.lower().endswith('.xlsx'):
        flash('El archivo debe tener formato .xlsx (Excel). Usa la plantilla descargable.', 'danger')
        return redirect(url_for('importar_alumnos'))

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb['Alumnos'] if 'Alumnos' in wb.sheetnames else wb.active
    except Exception:
        flash('No se pudo leer el archivo. Verifica que sea un .xlsx válido generado con la plantilla.', 'danger')
        return redirect(url_for('importar_alumnos'))

    encabezados = [(c.value.strip() if isinstance(c.value, str) else c.value) for c in ws[1]]
    nombres_columnas_conocidas = [nombre for nombre, _, _ in COLUMNAS_IMPORTACION_ALUMNOS]

    indice_columna = {
        nombre: encabezados.index(nombre)
        for nombre in nombres_columnas_conocidas
        if nombre in encabezados
    }

    faltantes = [
        nombre for nombre, requerido, _ in COLUMNAS_IMPORTACION_ALUMNOS
        if requerido and nombre not in indice_columna
    ]
    if faltantes:
        flash(
            f'Al archivo le faltan columnas obligatorias: {", ".join(faltantes)}. '
            'Descarga la plantilla más reciente y no cambies los nombres de las columnas.',
            'danger'
        )
        return redirect(url_for('importar_alumnos'))

    def valor_de(fila, nombre_col):
        idx = indice_columna.get(nombre_col)
        if idx is None or idx >= len(fila):
            return None
        valor = fila[idx].value
        if isinstance(valor, str):
            valor = valor.strip()
        return valor if valor not in ('', None) else None

    planes_por_clave = {
        p.clave_carrera: p for p in PlanEstudio.query.filter_by(activo=True).all()
    }

    exitosos = []
    errores = []

    for num_fila, fila in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(c.value in (None, '') for c in fila):
            continue  # Fila completamente vacía, se omite sin generar error

        fila_errores = []

        nombre_completo = valor_de(fila, 'nombre_completo')
        curp = str(valor_de(fila, 'curp') or '').upper()
        clave_carrera = str(valor_de(fila, 'clave_carrera') or '').upper()
        sexo = valor_de(fila, 'sexo')

        if not nombre_completo or len(str(nombre_completo)) < 5:
            fila_errores.append('nombre_completo inválido o vacío')

        if not re.match(r'^[A-Z0-9]{18}$', curp):
            fila_errores.append('CURP inválida (deben ser 18 caracteres)')
        elif Alumno.query.filter_by(curp=curp).first():
            fila_errores.append(f'ya existe un alumno con la CURP {curp}')

        plan = planes_por_clave.get(clave_carrera)
        if not plan:
            fila_errores.append(f'clave_carrera "{clave_carrera}" no corresponde a ningún plan activo')

        if not sexo:
            fila_errores.append('sexo vacío')

        def parsear_fecha(nombre_columna):
            crudo = valor_de(fila, nombre_columna)
            if isinstance(crudo, datetime):
                return crudo.date(), None
            if isinstance(crudo, str):
                try:
                    return datetime.strptime(crudo, '%Y-%m-%d').date(), None
                except ValueError:
                    return None, f'{nombre_columna} con formato inválido (usa AAAA-MM-DD)'
            return None, f'{nombre_columna} vacía'

        fecha_nacimiento, error_fn = parsear_fecha('fecha_nacimiento')
        if error_fn:
            fila_errores.append(error_fn)

        fecha_certificado, error_fc = parsear_fecha('fecha_certificado_prepa')
        if error_fc:
            fila_errores.append(error_fc)

        estatus_raw = str(valor_de(fila, 'estatus') or 'ACTIVO').upper()
        if estatus_raw not in EstatusAlumno.__members__:
            fila_errores.append(f'estatus "{estatus_raw}" no es válido')

        turno = None
        turno_raw = valor_de(fila, 'turno')
        if turno_raw:
            turno_raw = str(turno_raw).upper()
            if turno_raw in TurnoAlumno.__members__:
                turno = TurnoAlumno[turno_raw]
            else:
                fila_errores.append(f'turno "{turno_raw}" no es válido')

        modalidad = None
        modalidad_raw = valor_de(fila, 'modalidad')
        if modalidad_raw:
            modalidad_raw = str(modalidad_raw).upper()
            if modalidad_raw in ModalidadEstudio.__members__:
                modalidad = ModalidadEstudio[modalidad_raw]
            else:
                fila_errores.append(f'modalidad "{modalidad_raw}" no es válida')

        cuatrimestre_raw = valor_de(fila, 'cuatrimestre_actual')
        cuatrimestre_actual = 1
        if cuatrimestre_raw is not None:
            try:
                cuatrimestre_actual = int(cuatrimestre_raw)
            except (TypeError, ValueError):
                fila_errores.append('cuatrimestre_actual debe ser un número')

        if fila_errores:
            errores.append({
                'fila': num_fila,
                'nombre': nombre_completo or '(sin nombre)',
                'errores': fila_errores,
            })
            continue

        alumno, error_creacion = crear_alumno_generando_matricula(
            plan,
            nombre_completo=nombre_completo,
            curp=curp,
            fecha_nacimiento=fecha_nacimiento,
            fecha_certificado_prepa=fecha_certificado,
            estatus=EstatusAlumno[estatus_raw],
            cuatrimestre_actual=cuatrimestre_actual,
            correo=valor_de(fila, 'correo'),
            telefono=valor_de(fila, 'telefono'),
            telefono_movil=valor_de(fila, 'telefono_movil'),
            sexo=sexo,
            numero_identificacion=valor_de(fila, 'numero_identificacion'),
            estado_civil=valor_de(fila, 'estado_civil'),
            nacionalidad=valor_de(fila, 'nacionalidad') or 'Mexicana',
            tipo_sangre=valor_de(fila, 'tipo_sangre'),
            domicilio_calle_numero=valor_de(fila, 'domicilio_calle_numero'),
            domicilio_ciudad=valor_de(fila, 'domicilio_ciudad'),
            domicilio_cp=valor_de(fila, 'domicilio_cp'),
            domicilio_estado=valor_de(fila, 'domicilio_estado'),
            contacto_emergencia_nombre=valor_de(fila, 'contacto_emergencia_nombre'),
            contacto_emergencia_telefono=valor_de(fila, 'contacto_emergencia_telefono'),
            contacto_emergencia_parentesco=valor_de(fila, 'contacto_emergencia_parentesco'),
            turno=turno,
            modalidad=modalidad,
            grupo_actual=valor_de(fila, 'grupo_actual'),
        )

        if error_creacion:
            # commit por fila individual (ver crear_alumno_generando_matricula):
            # esta falla NO afecta a las filas anteriores, que ya quedaron
            # guardadas en la base de datos de forma independiente.
            errores.append({
                'fila': num_fila,
                'nombre': nombre_completo,
                'errores': [error_creacion],
            })
            continue

        exitosos.append({'fila': num_fila, 'nombre': nombre_completo, 'matricula': alumno.matricula_id})

    if exitosos:
        flash(f'Se importaron {len(exitosos)} alumno(s) correctamente.', 'success')

    if errores:
        flash(f'{len(errores)} fila(s) no se pudieron importar (ver detalle abajo).', 'warning')

    return render_template('importar_alumnos.html', exitosos=exitosos, errores=errores)


# ---------------------------------------------------------------------------
# MÓDULO DE DOCUMENTOS DEL EXPEDIENTE
# Permite completar los datos que NO se piden en el auto-registro público
# (escuela de prepa, alergias, teléfono del tutor) y subir los archivos
# digitalizados: comprobante de domicilio, INE y CURP.
# ---------------------------------------------------------------------------

def extension_permitida(nombre_archivo: str) -> bool:
    return (
        '.' in nombre_archivo
        and nombre_archivo.rsplit('.', 1)[1].lower() in app.config['EXTENSIONES_PERMITIDAS']
    )


# Mapeo de <name> del <input type="file"> -> TipoDocumento correspondiente
CAMPOS_DOCUMENTOS = {
    'archivo_domicilio': TipoDocumento.COMPROBANTE_DOMICILIO,
    'archivo_ine': TipoDocumento.INE,
    'archivo_curp': TipoDocumento.CURP_DOC,
    'archivo_acta': TipoDocumento.ACTA_NACIMIENTO,
    'archivo_certificado': TipoDocumento.CERTIFICADO_PREPA,
    'archivo_foto': TipoDocumento.FOTOGRAFIA,
}


@app.route('/alumno/<matricula>/documentos', methods=['GET', 'POST'])
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CAPTURADOR')
def documentos(matricula):
    alumno = Alumno.query.get_or_404(matricula)

    if request.method == 'POST':
        # --- 1. Actualizar datos complementarios del expediente ---
        alumno.escuela_prepa = request.form.get('escuela_prepa', '').strip() or alumno.escuela_prepa
        alumno.alergias_condiciones = request.form.get('alergias_condiciones', '').strip() or None
        alumno.telefono = request.form.get('telefono', '').strip() or alumno.telefono
        alumno.telefono_tutor = request.form.get('telefono_tutor', '').strip() or None

        # --- 1b. Seguimiento administrativo y académico ---
        cuatrimestre_raw = request.form.get('cuatrimestre_actual', '').strip()
        if cuatrimestre_raw.isdigit():
            alumno.cuatrimestre_actual = int(cuatrimestre_raw)
        alumno.documentacion_pendiente = request.form.get('documentacion_pendiente', '').strip() or None
        alumno.materias_adeudadas = request.form.get('materias_adeudadas', '').strip() or None
        alumno.faltas_administrativas = request.form.get('faltas_administrativas', '').strip() or None

        # --- 2. Procesar cada archivo subido (todos opcionales) ---
        carpeta_alumno = os.path.join(app.config['UPLOAD_FOLDER'], alumno.matricula_id)
        archivos_guardados = 0

        for campo_form, tipo_doc in CAMPOS_DOCUMENTOS.items():
            archivo = request.files.get(campo_form)

            if not archivo or archivo.filename == '':
                continue  # No se seleccionó archivo para este campo, se omite

            if not extension_permitida(archivo.filename):
                flash(
                    f'El archivo de "{tipo_doc.value}" tiene un formato no permitido '
                    f'(solo PDF, JPG o PNG).',
                    'danger'
                )
                continue

            os.makedirs(carpeta_alumno, exist_ok=True)
            nombre_seguro = secure_filename(archivo.filename)
            nombre_final = f'{tipo_doc.name}_{int(datetime.utcnow().timestamp())}_{nombre_seguro}'
            ruta_absoluta = os.path.join(carpeta_alumno, nombre_final)
            archivo.save(ruta_absoluta)

            documento = DocumentoAlumno(
                matricula_fk=alumno.matricula_id,
                tipo_documento=tipo_doc,
                nombre_archivo_original=archivo.filename,
                # OJO: aquí SIEMPRE forward-slash, aunque sea Windows, porque esta
                # ruta se usa para construir URLs (Flask sirve /static/... con '/').
                # os.path.join() usaría '\' en Windows y rompería el link (error 404).
                ruta_archivo=f'{alumno.matricula_id}/{nombre_final}',
            )
            db.session.add(documento)
            archivos_guardados += 1

        db.session.commit()

        if archivos_guardados:
            flash(f'Se guardaron {archivos_guardados} documento(s) y se actualizó la información.', 'success')
        else:
            flash('Se actualizó la información del expediente.', 'success')

        return redirect(url_for('documentos', matricula=matricula))

    documentos_alumno = (
        DocumentoAlumno.query
        .filter_by(matricula_fk=matricula)
        .order_by(DocumentoAlumno.fecha_subida.desc())
        .all()
    )
    return render_template(
        'documentos.html',
        alumno=alumno,
        documentos=documentos_alumno,
        max_cuatrimestres=app.config['CUATRIMESTRES_MAXIMOS']
    )


@app.route('/alumno/<matricula>/documento/<int:doc_id>/ver')
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CAPTURADOR')
def ver_documento(matricula, doc_id):
    """
    Sirve el archivo físico de un documento del expediente, exigiendo
    sesión activa + rol adecuado (a diferencia de servirlo directamente
    desde /static/, que no requiere ningún login).

    SECURITY-NOTE: verificamos explícitamente que el doc_id pedido
    pertenezca a LA MISMA matrícula que viene en la URL. Sin este chequeo,
    alguien con sesión válida pero de bajo rango podría cambiar el doc_id
    en la URL para intentar ver el documento de OTRO alumno cuya matrícula
    no conoce -- aunque ambos estén detrás del mismo control de rol, cada
    documento debe amarrarse a su propio expediente.
    """
    documento = DocumentoAlumno.query.get_or_404(doc_id)
    if documento.matricula_fk != matricula:
        abort(404)

    carpeta_absoluta = os.path.join(
        app.config['UPLOAD_FOLDER'],
        os.path.dirname(documento.ruta_archivo)
    )
    nombre_archivo = os.path.basename(documento.ruta_archivo)

    # send_from_directory ya protege internamente contra path traversal
    # (rutas tipo "../../etc/passwd"), pero igual construimos la ruta a
    # partir de datos que nosotros mismos generamos al subir el archivo
    # (ver documentos()), nunca a partir de un parámetro de la URL.
    return send_from_directory(carpeta_absoluta, nombre_archivo)


@app.route('/documento/<int:doc_id>/eliminar', methods=['POST'])
@rol_requerido('DIRECTIVO')
def eliminar_documento(doc_id):
    """
    Elimina un documento subido por equivocación: borra el archivo físico
    del disco y su registro en la base de datos, y regresa al expediente
    del mismo alumno.
    """
    documento = DocumentoAlumno.query.get_or_404(doc_id)
    matricula = documento.matricula_fk
    tipo_valor = documento.tipo_documento.value

    ruta_absoluta = os.path.join(
        app.config['UPLOAD_FOLDER'],
        documento.ruta_archivo.replace('/', os.sep)
    )

    try:
        if os.path.exists(ruta_absoluta):
            os.remove(ruta_absoluta)
    except OSError:
        # Si el archivo físico ya no está en disco, no bloqueamos el borrado
        # del registro en BD; igual se lo informamos al usuario.
        flash('El archivo físico ya no se encontró en el servidor; se eliminó el registro.', 'warning')

    db.session.delete(documento)
    db.session.commit()

    flash(f'Se eliminó el documento "{tipo_valor}".', 'success')
    return redirect(url_for('documentos', matricula=matricula))


# ---------------------------------------------------------------------------
# PERFIL DEL ESTUDIANTE / EXPEDIENTE ACADÉMICO
# ---------------------------------------------------------------------------

@app.route('/alumno/<matricula>/expediente')
@login_required
def ver_expediente(matricula):
    alumno = Alumno.query.get_or_404(matricula)

    # Todas las materias del plan del alumno (el "Escudo": solo las de SU plan)
    materias_plan = (
        Materia.query
        .filter_by(id_plan_fk=alumno.id_plan_fk)
        .order_by(Materia.cuatrimestre.asc(), Materia.nombre.asc())
        .all()
    )

    # Última calificación registrada por materia (si una materia se recursó,
    # más de una Calificacion puede existir; nos quedamos con la más reciente)
    calif_por_materia = {}
    for c in sorted(alumno.calificaciones, key=lambda c: c.fecha_captura):
        calif_por_materia[c.id_materia_fk] = c

    # Agrupar por cuatrimestre para el acordeón
    cuatrimestres = {}
    for materia in materias_plan:
        cuatrimestres.setdefault(materia.cuatrimestre, []).append({
            'materia': materia,
            'calificacion': calif_por_materia.get(materia.id)
        })

    # Materias del plan que aún no están aprobadas (calificación >= 6).
    # Se usa para advertir (sin bloquear) al intentar marcar al alumno como Egresado.
    materias_no_aprobadas = [
        materia for materia in materias_plan
        if not (calif_por_materia.get(materia.id) and calif_por_materia[materia.id].calificacion_final >= 6)
    ]

    return render_template(
        'expediente.html',
        alumno=alumno,
        cuatrimestres=cuatrimestres,
        max_cuatrimestres=app.config['CUATRIMESTRES_MAXIMOS'],
        estatus_disponibles=list(EstatusAlumno),
        materias_no_aprobadas=materias_no_aprobadas,
        historial_estatus=alumno.historial_estatus
    )


@app.route('/alumno/<matricula>/ficha')
@login_required
def ficha_inscripcion(matricula):
    """
    Ficha de Inscripción imprimible (frente + reverso), a partir de los
    datos ya capturados en el registro público y el módulo de documentos.
    No requiere formulario propio: es una vista de solo lectura para
    imprimir/archivar, con checklist de documentos recibidos en el reverso.
    """
    alumno = Alumno.query.get_or_404(matricula)

    documentos_alumno = DocumentoAlumno.query.filter_by(matricula_fk=matricula).all()
    documentos_subidos = {doc.tipo_documento for doc in documentos_alumno}

    return render_template(
        'ficha_inscripcion.html',
        alumno=alumno,
        tipos_documento=list(TipoDocumento),
        documentos_subidos=documentos_subidos
    )


@app.route('/alumno/<matricula>/cambiar-estatus', methods=['POST'])
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CAPTURADOR')
def cambiar_estatus(matricula):
    """
    Cambia el estatus del alumno (ej. Pendiente -> Activo -> Egresado).
    Registra el cambio en HistorialEstatus para auditoría. Si se intenta
    marcar como EGRESADO y el alumno tiene materias sin aprobar, se pide
    un comentario justificando el motivo (advertencia, no bloqueo total).
    """
    alumno = Alumno.query.get_or_404(matricula)
    nuevo_estatus_raw = request.form.get('nuevo_estatus', '')
    comentario = request.form.get('comentario', '').strip() or None

    if nuevo_estatus_raw not in EstatusAlumno.__members__:
        flash('El estatus indicado no es válido.', 'danger')
        return redirect(url_for('ver_expediente', matricula=matricula))

    nuevo_estatus = EstatusAlumno[nuevo_estatus_raw]

    if nuevo_estatus == alumno.estatus:
        flash('El alumno ya tiene ese estatus; no se realizó ningún cambio.', 'warning')
        return redirect(url_for('ver_expediente', matricula=matricula))

    if nuevo_estatus == EstatusAlumno.EGRESADO:
        materias_plan = Materia.query.filter_by(id_plan_fk=alumno.id_plan_fk).all()
        calif_por_materia = {c.id_materia_fk: c for c in alumno.calificaciones}
        faltantes = [
            m for m in materias_plan
            if not (calif_por_materia.get(m.id) and calif_por_materia[m.id].calificacion_final >= 6)
        ]

        if faltantes and not comentario:
            nombres = ', '.join(m.nombre for m in faltantes[:5])
            extra = f' y {len(faltantes) - 5} más' if len(faltantes) > 5 else ''
            flash(
                f'El alumno tiene {len(faltantes)} materia(s) sin aprobar ({nombres}{extra}). '
                'Si de verdad quieres marcarlo como Egresado, agrega un comentario '
                'explicando el motivo y vuelve a intentarlo.',
                'warning'
            )
            return redirect(url_for('ver_expediente', matricula=matricula))

        if alumno.tiene_adeudo() and not comentario:
            saldo = alumno.saldo_total_adeudado()
            flash(
                f'El alumno tiene un adeudo económico de ${saldo:.2f}. '
                'Si de verdad quieres marcarlo como Egresado, agrega un comentario '
                'explicando el motivo (ver Cobros para el detalle) y vuelve a intentarlo.',
                'warning'
            )
            return redirect(url_for('ver_expediente', matricula=matricula))

    registro = HistorialEstatus(
        matricula_fk=alumno.matricula_id,
        usuario_fk=current_user.id,
        estatus_anterior=alumno.estatus,
        estatus_nuevo=nuevo_estatus,
        comentario=comentario,
    )
    alumno.estatus = nuevo_estatus

    if nuevo_estatus == EstatusAlumno.ACTIVO and not alumno.fecha_validacion:
        alumno.fecha_validacion = datetime.utcnow()

    db.session.add(registro)
    db.session.commit()

    flash(f'Estatus del alumno actualizado a "{nuevo_estatus.value}".', 'success')
    return redirect(url_for('ver_expediente', matricula=matricula))


# ---------------------------------------------------------------------------
# SISTEMA DE COBROS
# Ver y consultar: ambos roles. Crear/cancelar cargos: solo Directivo
# (define la estructura financiera). Registrar un pago: ambos roles (es
# el trabajo diario de Control Escolar en la ventanilla — "actualizar").
# ---------------------------------------------------------------------------

@app.route('/alumno/<matricula>/cobros')
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CONTADOR')
def cobros(matricula):
    alumno = Alumno.query.get_or_404(matricula)
    cargos = Cargo.query.filter_by(matricula_fk=matricula).order_by(Cargo.fecha_generacion.desc()).all()

    # Recargos "automáticos": se recalculan cada vez que se consulta la
    # pantalla, usando la configuración VIGENTE (auto-ajustable). No
    # depende de ningún cron job en segundo plano.
    hubo_cambios = False
    for cargo in cargos:
        recargo_antes = cargo.recargo_aplicado
        cargo.actualizar_recargo_si_vencido()
        if cargo.recargo_aplicado != recargo_antes:
            cargo.actualizar_estatus()
            hubo_cambios = True
    if hubo_cambios:
        db.session.commit()

    total_adeudado = sum(
        (c.saldo_pendiente() for c in cargos if c.estatus != EstatusCargo.CANCELADO),
        Decimal('0.00')
    )

    return render_template(
        'cobros.html',
        alumno=alumno,
        cargos=cargos,
        total_adeudado=total_adeudado,
        metodos_pago=list(MetodoPago),
        conceptos_cobro=ConceptoCobro.query.filter_by(activo=True).order_by(ConceptoCobro.nombre.asc()).all(),
    )


@app.route('/alumno/<matricula>/cobros/nuevo', methods=['POST'])
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CONTADOR')
def nuevo_cargo(matricula):
    alumno = Alumno.query.get_or_404(matricula)

    concepto_cobro_id_raw = request.form.get('concepto_cobro_id', '').strip()
    monto_raw = request.form.get('monto', '').strip()
    periodo_escolar = request.form.get('periodo_escolar', '').strip() or None
    fecha_vencimiento_raw = request.form.get('fecha_vencimiento', '').strip()

    errores = []

    concepto_cobro = None
    if not concepto_cobro_id_raw:
        errores.append('Selecciona un concepto del catálogo.')
    else:
        try:
            concepto_cobro = ConceptoCobro.query.get(int(concepto_cobro_id_raw))
        except (ValueError, TypeError):
            concepto_cobro = None
        if not concepto_cobro or not concepto_cobro.activo:
            errores.append('El concepto seleccionado no es válido.')

    monto = None
    try:
        monto = Decimal(monto_raw)
        if monto <= 0:
            errores.append('El monto debe ser mayor a 0.')
    except InvalidOperation:
        errores.append('El monto no es un número válido.')

    fecha_vencimiento = None
    if fecha_vencimiento_raw:
        try:
            fecha_vencimiento = datetime.strptime(fecha_vencimiento_raw, '%Y-%m-%d').date()
        except ValueError:
            errores.append('La fecha de vencimiento no es válida.')

    if errores:
        for error in errores:
            flash(error, 'danger')
        return redirect(url_for('cobros', matricula=matricula))

    nuevo = Cargo(
        matricula_fk=alumno.matricula_id,
        concepto_cobro_fk=concepto_cobro.id,
        concepto=concepto_cobro.nombre,  # Denormalizado para mostrar sin necesidad de join
        monto=monto,
        periodo_escolar=periodo_escolar,
        fecha_vencimiento=fecha_vencimiento,
        generado_por_fk=current_user.id,
    )
    db.session.add(nuevo)
    db.session.commit()

    flash(f'Cargo "{concepto_cobro.nombre}" agregado correctamente.', 'success')
    return redirect(url_for('cobros', matricula=matricula))


def enviar_comprobante_pago(alumno, cargo, pago):
    """
    Envía el comprobante de pago al correo del alumno. Si el alumno no
    tiene correo registrado, o si falla el envío (sin internet, SMTP
    caído, etc.), NO debe romper el registro del pago -- el pago ya
    quedó guardado en la BD; solo se avisa al usuario que el correo no
    se pudo mandar.
    """
    if not alumno.correo:
        return False, 'El alumno no tiene correo registrado.'

    try:
        mensaje = Message(
            subject=f'Comprobante de Pago - Folio {pago.folio or ("#" + str(pago.id))}',
            recipients=[alumno.correo],
            body=(
                f'Hola {alumno.nombre_completo},\n\n'
                f'Se registró tu pago con los siguientes datos:\n\n'
                f'Concepto: {cargo.concepto}\n'
                f'Monto pagado: ${pago.monto_pagado}\n'
                f'Fecha: {pago.fecha_pago.strftime("%d/%m/%Y %H:%M")}\n'
                f'Folio: {pago.folio or ("#" + str(pago.id))}\n'
                f'Saldo pendiente del cargo: ${cargo.saldo_pendiente()}\n\n'
                f'Este es un correo automático del Sistema de Gestión Escolar.'
            ),
        )
        mail.send(mensaje)
        return True, None
    except Exception as error:
        return False, str(error)


@app.route('/cobro/<int:cargo_id>/pagar', methods=['POST'])
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CONTADOR')
def registrar_pago(cargo_id):
    cargo = Cargo.query.get_or_404(cargo_id)

    if cargo.estatus == EstatusCargo.CANCELADO:
        flash('Este cargo está cancelado; no se le pueden registrar pagos.', 'danger')
        return redirect(url_for('cobros', matricula=cargo.matricula_fk))

    monto_raw = request.form.get('monto_pagado', '').strip()
    metodo_raw = request.form.get('metodo_pago', 'EFECTIVO').upper()
    referencia = request.form.get('referencia', '').strip() or None
    comentario = request.form.get('comentario', '').strip() or None

    try:
        monto_pagado = Decimal(monto_raw)
        if monto_pagado <= 0:
            raise InvalidOperation()
    except InvalidOperation:
        flash('El monto pagado no es válido.', 'danger')
        return redirect(url_for('cobros', matricula=cargo.matricula_fk))

    saldo = cargo.saldo_pendiente()
    if monto_pagado > saldo:
        flash(
            f'El monto pagado (${monto_pagado}) es mayor al saldo pendiente (${saldo}). '
            'Verifica el monto.',
            'danger'
        )
        return redirect(url_for('cobros', matricula=cargo.matricula_fk))

    if metodo_raw not in MetodoPago.__members__:
        metodo_raw = 'EFECTIVO'

    pago = Pago(
        cargo=cargo,
        monto_pagado=monto_pagado,
        metodo_pago=MetodoPago[metodo_raw],
        referencia=referencia,
        capturado_por_fk=current_user.id,
        comentario=comentario,
    )
    db.session.add(pago)
    db.session.flush()  # Asigna pago.id (lo necesitamos para armar el folio) y refleja el pago en saldo_pendiente()

    pago.folio = f'PAGO-{datetime.utcnow().year}-{pago.id:06d}'

    cargo.actualizar_estatus()
    db.session.commit()

    enviado, error_correo = enviar_comprobante_pago(cargo.alumno, cargo, pago)

    flash(f'Pago de ${monto_pagado} registrado correctamente. Folio: {pago.folio}', 'success')
    if enviado:
        flash(f'Comprobante enviado a {cargo.alumno.correo}.', 'success')
    else:
        flash(f'El pago se guardó bien, pero no se pudo enviar el comprobante por correo ({error_correo}).', 'warning')

    return redirect(url_for('cobros', matricula=cargo.matricula_fk))


@app.route('/cobro/<int:cargo_id>/cancelar', methods=['POST'])
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CONTADOR')
def cancelar_cargo(cargo_id):
    cargo = Cargo.query.get_or_404(cargo_id)
    comentario = request.form.get('comentario', '').strip() or None

    cargo.estatus = EstatusCargo.CANCELADO
    cargo.comentario = comentario
    db.session.commit()

    flash(f'Cargo "{cargo.concepto}" cancelado.', 'success')
    return redirect(url_for('cobros', matricula=cargo.matricula_fk))


@app.route('/pago/<int:pago_id>/recibo')
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CONTADOR')
def recibo_pago(pago_id):
    pago = Pago.query.get_or_404(pago_id)
    return render_template('recibo_pago.html', pago=pago, cargo=pago.cargo, alumno=pago.cargo.alumno)


@app.route('/alumno/<matricula>/estado-cuenta')
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CONTADOR')
def estado_cuenta(matricula):
    """
    "Tira de pagos" del alumno durante todo su ciclo escolar: histórico
    completo de cargos y pagos, imprimible. A diferencia de /cobros (que
    es la pantalla de trabajo diario), esta vista es de solo lectura,
    pensada para entregarse o archivarse.
    """
    alumno = Alumno.query.get_or_404(matricula)
    cargos = Cargo.query.filter_by(matricula_fk=matricula).order_by(Cargo.fecha_generacion.asc()).all()

    for cargo in cargos:
        cargo.actualizar_recargo_si_vencido()
    db.session.commit()

    total_cargado = sum((c.monto + c.recargo_aplicado for c in cargos), Decimal('0.00'))
    total_pagado = sum((c.total_pagado() for c in cargos), Decimal('0.00'))
    saldo_total = alumno.saldo_total_adeudado()

    return render_template(
        'estado_cuenta.html',
        alumno=alumno,
        cargos=cargos,
        total_cargado=total_cargado,
        total_pagado=total_pagado,
        saldo_total=saldo_total,
    )


@app.route('/reportes/cobros-del-dia')
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CONTADOR')
def reporte_cobros_del_dia():
    """
    Reporte a demanda de todos los pagos capturados en una fecha (por
    defecto, hoy). Cualquiera de los dos roles puede consultarlo —
    es información, no una acción de modificación.
    """
    fecha_raw = request.args.get('fecha', '')
    try:
        fecha_reporte = datetime.strptime(fecha_raw, '%Y-%m-%d').date() if fecha_raw else datetime.utcnow().date()
    except ValueError:
        fecha_reporte = datetime.utcnow().date()
        flash('La fecha indicada no era válida; se muestra el día de hoy.', 'warning')

    inicio_dia = datetime.combine(fecha_reporte, datetime.min.time())
    fin_dia = datetime.combine(fecha_reporte, datetime.max.time())

    pagos_del_dia = (
        Pago.query
        .filter(Pago.fecha_pago >= inicio_dia, Pago.fecha_pago <= fin_dia)
        .order_by(Pago.fecha_pago.asc())
        .all()
    )

    total_del_dia = sum((p.monto_pagado for p in pagos_del_dia), Decimal('0.00'))

    totales_por_concepto = {}
    totales_por_metodo = {}
    for pago in pagos_del_dia:
        concepto = pago.cargo.concepto if pago.cargo else 'Sin concepto'
        totales_por_concepto[concepto] = totales_por_concepto.get(concepto, Decimal('0.00')) + pago.monto_pagado
        totales_por_metodo[pago.metodo_pago.value] = totales_por_metodo.get(pago.metodo_pago.value, Decimal('0.00')) + pago.monto_pagado

    return render_template(
        'reporte_cobros_dia.html',
        fecha_reporte=fecha_reporte,
        pagos_del_dia=pagos_del_dia,
        total_del_dia=total_del_dia,
        totales_por_concepto=totales_por_concepto,
        totales_por_metodo=totales_por_metodo,
    )


@app.route('/planes/mensualidades', methods=['GET', 'POST'])
@rol_requerido('DIRECTIVO')
def planes_mensualidades():
    """
    Pantalla mínima para que Dirección ajuste el precio de mensualidad de
    cada carrera (cada una puede costar distinto). Solo edita ese campo;
    el CRUD completo de Planes de Estudio sigue pendiente como tarea aparte.
    """
    if request.method == 'POST':
        plan_id = request.form.get('plan_id', '').strip()
        monto_raw = request.form.get('monto_mensualidad', '').strip()
        plan = PlanEstudio.query.get_or_404(int(plan_id)) if plan_id.isdigit() else None

        if not plan:
            flash('Plan de estudios no encontrado.', 'danger')
        else:
            try:
                monto = Decimal(monto_raw)
                if monto < 0:
                    raise InvalidOperation
                plan.monto_mensualidad = monto
                db.session.commit()
                flash(f'Mensualidad de "{plan.nombre}" actualizada a ${monto}.', 'success')
            except InvalidOperation:
                flash('El monto no es un número válido.', 'danger')

        return redirect(url_for('planes_mensualidades'))

    planes = PlanEstudio.query.order_by(PlanEstudio.nombre.asc()).all()
    return render_template('planes_mensualidades.html', planes=planes)


@app.route('/conceptos-cobro', methods=['GET', 'POST'])
@rol_requerido('DIRECTIVO', 'CONTADOR')
def conceptos_cobro():
    """Catálogo de conceptos de cobro — se administra aquí, NUNCA como texto libre al capturar un cargo."""
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()

        if len(nombre) < 3:
            flash('El nombre del concepto debe tener al menos 3 caracteres.', 'danger')
        elif ConceptoCobro.query.filter_by(nombre=nombre).first():
            flash(f'Ya existe un concepto llamado "{nombre}".', 'danger')
        else:
            nuevo = ConceptoCobro(nombre=nombre, activo=True)
            db.session.add(nuevo)
            db.session.commit()
            flash(f'Concepto "{nombre}" agregado al catálogo.', 'success')

        return redirect(url_for('conceptos_cobro'))

    conceptos = ConceptoCobro.query.order_by(ConceptoCobro.activo.desc(), ConceptoCobro.nombre.asc()).all()
    return render_template('conceptos_cobro.html', conceptos=conceptos)


@app.route('/conceptos-cobro/<int:concepto_id>/toggle', methods=['POST'])
@rol_requerido('DIRECTIVO', 'CONTADOR')
def toggle_concepto_cobro(concepto_id):
    concepto = ConceptoCobro.query.get_or_404(concepto_id)
    concepto.activo = not concepto.activo
    db.session.commit()

    estado = 'activado' if concepto.activo else 'desactivado'
    flash(f'El concepto "{concepto.nombre}" fue {estado}.', 'success')
    return redirect(url_for('conceptos_cobro'))


@app.route('/configuracion/cobros', methods=['GET', 'POST'])
@rol_requerido('DIRECTIVO', 'CONTADOR')
def configuracion_cobros():
    """
    Configuración de recargos por atraso — auto-ajustable: cada
    universidad define su propia fórmula aquí, sin tocar código.
    """
    config = ConfiguracionCobros.obtener()

    if request.method == 'POST':
        tipo_raw = request.form.get('tipo_recargo', '')
        valor_raw = request.form.get('valor_recargo', '').strip()
        dias_gracia_raw = request.form.get('dias_gracia', '0').strip()

        errores = []

        if tipo_raw not in TipoRecargo.__members__:
            errores.append('Selecciona un tipo de recargo válido.')

        valor = None
        try:
            valor = Decimal(valor_raw)
            if valor < 0:
                errores.append('El valor del recargo no puede ser negativo.')
        except InvalidOperation:
            errores.append('El valor del recargo no es un número válido.')

        try:
            dias_gracia = int(dias_gracia_raw)
            if dias_gracia < 0:
                errores.append('Los días de gracia no pueden ser negativos.')
        except ValueError:
            errores.append('Los días de gracia deben ser un número entero.')
            dias_gracia = 0

        if errores:
            for error in errores:
                flash(error, 'danger')
            return redirect(url_for('configuracion_cobros'))

        config.tipo_recargo = TipoRecargo[tipo_raw]
        config.valor_recargo = valor
        config.dias_gracia = dias_gracia
        db.session.commit()

        flash('Configuración de recargos actualizada correctamente.', 'success')
        return redirect(url_for('configuracion_cobros'))

    return render_template('configuracion_cobros.html', config=config, tipos_recargo=list(TipoRecargo))


# ---------------------------------------------------------------------------
# MÓDULO DE CAPTURA DE CALIFICACIONES (BOLETA)
# Solo Control Escolar captura, con base en actas físicas firmadas.
# El "Escudo" se aplica filtrando SIEMPRE por id_plan_fk del alumno.
# ---------------------------------------------------------------------------

@app.route('/alumno/<matricula>/boleta', methods=['GET', 'POST'])
@rol_requerido('DIRECTIVO', 'ADMINISTRATIVO', 'CAPTURADOR')
def boleta(matricula):
    alumno = Alumno.query.get_or_404(matricula)

    cuatrimestre_seleccionado = request.values.get('cuatrimestre', type=int)
    if not cuatrimestre_seleccionado:
        cuatrimestre_seleccionado = alumno.cuatrimestre_actual or 1

    # ESCUDO DEL PLAN DE ESTUDIOS: solo materias de ESTE plan y ESTE cuatrimestre.
    # Nunca se ofrece (ni se acepta) una materia fuera de esta consulta.
    materias = (
        Materia.query
        .filter_by(id_plan_fk=alumno.id_plan_fk, cuatrimestre=cuatrimestre_seleccionado)
        .order_by(Materia.nombre.asc())
        .all()
    )

    if request.method == 'POST':
        periodo_escolar = request.form.get('periodo_escolar', '').strip()
        numero_acta = request.form.get('numero_acta', '').strip() or None
        # SECURITY-NOTE: antes venía de un <input> de texto libre en el
        # formulario, así que cualquiera podía escribir el nombre de otra
        # persona ahí -- el registro de auditoría no era confiable. Ahora
        # se toma directo de la sesión autenticada, igual que ya se hacía
        # en Cargo.generado_por_fk y Pago.capturado_por_fk.
        capturado_por = current_user.nombre_completo

        if not periodo_escolar:
            flash('Indica el periodo escolar (ej. "2026-A") antes de guardar.', 'danger')
            return redirect(url_for('boleta', matricula=matricula, cuatrimestre=cuatrimestre_seleccionado))

        guardadas = 0
        for materia in materias:
            valor_raw = request.form.get(f'calificacion_{materia.id}', '').strip()

            if valor_raw == '':
                continue  # Casilla vacía = aún no se captura, se omite sin error

            try:
                valor = float(valor_raw)
            except ValueError:
                flash(f'La calificación de "{materia.nombre}" no es un número válido.', 'danger')
                continue

            if valor < 0 or valor > 10:
                flash(f'La calificación de "{materia.nombre}" debe estar entre 0 y 10.', 'danger')
                continue

            # Doble verificación del "Escudo" a nivel de objeto, por si acaso.
            if not alumno.materia_pertenece_a_su_plan(materia):
                flash('Se rechazó una materia que no pertenece al plan del alumno.', 'danger')
                continue

            existente = Calificacion.query.filter_by(
                matricula_fk=alumno.matricula_id,
                id_materia_fk=materia.id,
                periodo_escolar=periodo_escolar
            ).first()

            if existente:
                existente.calificacion_final = valor
                existente.numero_acta = numero_acta
                existente.capturado_por = capturado_por
            else:
                db.session.add(Calificacion(
                    matricula_fk=alumno.matricula_id,
                    id_materia_fk=materia.id,
                    calificacion_final=valor,
                    periodo_escolar=periodo_escolar,
                    numero_acta=numero_acta,
                    capturado_por=capturado_por,
                ))
            guardadas += 1

        db.session.commit()
        flash(f'Se guardaron {guardadas} calificación(es) del {cuatrimestre_seleccionado}° cuatrimestre.', 'success')
        return redirect(url_for('boleta', matricula=matricula, cuatrimestre=cuatrimestre_seleccionado))

    # Precargar calificaciones existentes de las materias de este cuatrimestre
    calificaciones_existentes = {
        c.id_materia_fk: c
        for c in alumno.calificaciones
        if c.materia.cuatrimestre == cuatrimestre_seleccionado
    }

    return render_template(
        'boleta.html',
        alumno=alumno,
        materias=materias,
        cuatrimestre_seleccionado=cuatrimestre_seleccionado,
        max_cuatrimestres=app.config['CUATRIMESTRES_MAXIMOS'],
        calificaciones_existentes=calificaciones_existentes,
    )


if __name__ == '__main__':
    with app.app_context():
        # Crea las tablas si no existen (solo desarrollo).
        # En producción se recomienda usar Flask-Migrate (ver Paso posterior).
        db.create_all()
    app.run(debug=True)
